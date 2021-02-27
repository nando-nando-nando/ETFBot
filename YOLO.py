# https://docs.tweepy.org/
import tweepy
import datetime
import urllib
from openpyxl import load_workbook


auth = tweepy.OAuthHandler('', '')
auth.set_access_token('', '')

# Auth with Twitter
try:
    api = tweepy.API(auth)
    if api.verify_credentials() == False:
        raise Exception("Couldn't authenticate with Twitter.")
except:
    print("Couldn't authenticate with Twitter.")
    exit()


## First, get the newest and previous workbooks. Find the differences. 

# Filenames
today = datetime.datetime.now().strftime("%m-%d-%y")
yesterday = (datetime.datetime.now() - datetime.timedelta(1)).strftime("%m-%d-%y")
fileLocNew = f"holdings/yolo/{today}.xlsx"
fileLocOld = f"holdings/yolo/{yesterday}.xlsx"
imgFileLocNew = f"holdings/yolo/imgs/AdvisorShares_YOLO_Holdings_{today}.png"

# Collect Excel sheet
dls = "https://advisorshares.com/wp-content/uploads/csv/holdings/AdvisorShares_YOLO_Holdings_File.xlsx"
urllib.request.urlretrieve(dls, fileLocNew)  # For Python 3

wbNew = load_workbook(fileLocNew, data_only=True)
wbOld = load_workbook(fileLocOld,  data_only=True)

sheetNew = wbNew.active
sheetOld = wbOld.active

for column_cells in sheetNew.columns: 
  unmerged_cells = list(filter(lambda cell_to_check: cell_to_check.coordinate not in sheetNew.merged_cells, column_cells)) 
  length = max(len(str(cell.value)) for cell in unmerged_cells) 
  sheetNew.column_dimensions[unmerged_cells[0].column_letter].width = length * .8
wbNew.save(fileLocNew)

#Convert Excel sheet to img for tweet attachment
import excel2img
excel2img.export_img(fileLocNew,imgFileLocNew,None, None)

# Setup dictionaries with old and new sets of ticker/share# pairs
newPairs = dict()
oldPairs = dict()
diffList = dict()
openedList = dict()
closedList = dict()
try:
    #Todo: make this all suck less
    row = 1
    for cell in sheetNew['C']:
        if  not cell.value or len(cell.value) < 2 or row == 1:
            row = row+1
            pass
        else:
            newPairs[cell.value] = sheetNew[f'F{row}'].value
            # print(f'{cell.value}:{newPairs[cell.value]}')
            row = row+1

    row = 1
    for cell in sheetOld['C']:
        if  not cell.value or len(cell.value) < 2 or row == 1:
            row = row+1
            pass
        else:
            oldPairs[cell.value] = sheetOld[f'F{row}'].value
            # print(f'{cell.value}:{oldPairs[cell.value]}')
            row = row+1

    #Separate out newly opened, closed, and updated positions
    for ticker in newPairs:
        if ticker in oldPairs:
            diff = (int(float(newPairs[ticker])) - int(float(oldPairs[ticker])))
            if diff != 0:
                diffList[ticker] = diff
        else: 
            openedList[ticker] = newPairs[ticker]
            
    for ticker in oldPairs:
        if ticker not in newPairs:
            closedList[ticker] = oldPairs[ticker]

    if not bool(diffList) and not bool(openedList) and not bool(closedList):
        print(f"There was no difference between the holdings for {today} and {yesterday}. No tweets today")
        exit()

    # Build the tweet message, paginate if necessary
    tweet = [f'The latest @AdvisorShares $YOLO holdings are out🌿🇨🇦\n#potstocks\n{today}\n\n']
    page = 0
    if bool(diffList):
        tweet[page] = tweet[page] + 'Position Changes\n' 
        for ticker in diffList:
            if len(tweet[page]) >= 240:
                page = page + 1
                tweet.append('')
            tweet[page] = tweet[page] + f'\n  ${ticker.ljust(8)}' + f'{diffList[ticker]:,d}'

    if bool(openedList):
        if len(tweet[page]) >= 240:
            page = page + 1
            tweet.append('')
            tweet[page] = tweet[page] + 'Opened Positions\n'
        else:
            tweet[page] = tweet[page] + '\n\nOpened Positions\n' 
        for ticker in openedList:
            if len(tweet[page]) >= 260:
                page = page + 1
                tweet.append('')
            tweet[page] = tweet[page] + f'\n  ${ticker.ljust(8)}' + f'{openedList[ticker]:,d}'

    if bool(closedList):
        if len(tweet[page]) >= 240:
            page = page + 1
            tweet.append('')
            tweet[page] = tweet[page] + 'Closed Positions\n' 
        else:
            tweet[page] = tweet[page] + '\n\nClosed Positions\n'
        for ticker in closedList:
            if len(tweet[page]) >= 260:
                page = page + 1
                tweet.append('')
            tweet[page] = tweet[page] + f'\n  ${ticker.ljust(8)}' + f'{closedList[ticker]:,d}'

    # Add a page number if there are more than one pages (280 chars max per tweet)
    if (page>0):
        pageCount = 0
        while pageCount <= page:
            tweet[pageCount] = tweet[pageCount] + f'\n\n{pageCount+1}/{page+1}' 
            print(tweet[pageCount])
            pageCount = pageCount + 1
    else:
        print(tweet[0])
except Exception as e:
    print("ERROR during processing, closing...")
    print(e.__dict__)
    exit()        
# #Upload image
# #Model: {'_api': <tweepy.api.API object at 0x0000025B82EB60A0>, 'media_id': 1362203886765838338, 'media_id_string': '1362203886765838338', 'size': 42703, 'expires_after_secs': 86400, 'image': {'image_type': 'image/png', 'w': 1007, 'h': 562}}
# mediaobj = api.media_upload(imgFileLocNew)

# #Send tweet(s)
# replyId = 0
# for pg in tweet:
#     try:
#         if replyId == 0:
#             status = api.update_status(status=pg, media_ids=[mediaobj.media_id_string])
#             replyId = status.id_str
#         else:
#             status = api.update_status(status=pg, in_reply_to_status_id=status.id_str)
#             replyId = status.id_str
#     except tweepy.TweepError as e:
#         print("ERROR during tweet, closing...")
#         print(e.__dict__)
#         print(e.api_code) #code 186 is tweet too long
#         exit()
#     except Exception as e:
#         print("ERROR during tweet, closing...")
#         print(e.__dict__)
#         exit()